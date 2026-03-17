"""
Excel structure analyzer using openpyxl.

Detects title rows, header rows, column-numbering rows, merged cells,
footer/footnote rows, and data boundaries in complex Excel files
BEFORE pandas loads them.  This avoids the common pitfall of pandas
flattening merged cells into NaN-filled rows that confuse downstream
header detection and LLM analysis.

Also detects multi-sheet workbooks and vertically stacked sub-tables
within a single sheet (e.g. same columns repeated under different
"Base Year" banners or "Constant Prices" / "Current Prices" sections).
"""

import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.logger import logger


@dataclass
class SubTable:
    """One logical sub-table found inside a sheet."""
    sheet_name: str
    header_rows: List[int]            # 1-indexed
    col_number_row: Optional[int]
    data_start_row: int               # 1-indexed
    data_end_row: int                 # 1-indexed
    min_data_col: int
    max_data_col: int
    num_header_rows: int
    merged_header_values: Optional[List[List[Optional[str]]]] = None
    label: Optional[str] = None       # e.g. "Constant Prices", "Current Prices"
    base_year: Optional[str] = None   # e.g. "2011-12", "2004-05"
    unit_hint: Optional[str] = None
    horizontal_split: bool = False    # True when split from a horizontally-tiled table
    label_column: Optional[str] = None  # column name for the label (e.g. "exchange")


@dataclass
class WorkbookStructure:
    """Result of full-workbook analysis."""
    title_hint: Optional[str] = None
    unit_hint: Optional[str] = None
    sheet_names: List[str] = field(default_factory=list)
    sub_tables: List[SubTable] = field(default_factory=list)
    # True when all sub-tables share the same header columns and
    # should be concatenated (with distinguishing columns) into one table.
    mergeable: bool = False


class ExcelStructure:
    """Result of Excel structure analysis."""

    def __init__(
        self,
        *,
        skip_rows: int = 0,
        header_rows: List[int],
        num_header_rows: int = 1,
        col_number_row: Optional[int] = None,
        data_start_row: int,
        data_end_row: int,
        footer_start_row: Optional[int] = None,
        min_data_col: int = 1,
        max_data_col: int = 1,
        unit_hint: Optional[str] = None,
        title_hint: Optional[str] = None,
        merged_header_values: Optional[List[List[Optional[str]]]] = None,
    ):
        self.skip_rows = skip_rows
        self.header_rows = header_rows            # 1-indexed Excel row numbers
        self.num_header_rows = num_header_rows
        self.col_number_row = col_number_row      # 1-indexed, or None
        self.data_start_row = data_start_row      # 1-indexed, first actual data row
        self.data_end_row = data_end_row          # 1-indexed, last actual data row
        self.footer_start_row = footer_start_row
        self.min_data_col = min_data_col          # 1-indexed
        self.max_data_col = max_data_col          # 1-indexed
        self.unit_hint = unit_hint                # e.g. "₹ Crores"
        self.title_hint = title_hint              # extracted title text
        self.merged_header_values = merged_header_values  # resolved header grid

    def to_dict(self) -> Dict:
        return {
            "skip_rows": self.skip_rows,
            "header_rows": self.header_rows,
            "num_header_rows": self.num_header_rows,
            "col_number_row": self.col_number_row,
            "data_start_row": self.data_start_row,
            "data_end_row": self.data_end_row,
            "footer_start_row": self.footer_start_row,
            "min_data_col": self.min_data_col,
            "max_data_col": self.max_data_col,
            "unit_hint": self.unit_hint,
            "title_hint": self.title_hint,
        }


# ── Patterns ──────────────────────────────────────────────────────────

_FOOTER_PATTERNS = re.compile(
    r"^\s*(notes?\s*:|source\s*:|"
    r"\*|@|#|†|‡|§|"
    r"\d+\.\s+[A-Z]|"                       # "1. Data for ..."
    r"p\s*[-–:]\s*provisional|"
    r"re\s*[-–:]\s*revised|"
    r"be\s*[-–:]\s*budget)",
    re.IGNORECASE,
)

_UNIT_PATTERNS = re.compile(
    r"\(?\s*₹\s*\w+\s*\)?|"
    r"\(\s*(?:Rs|INR|rupee|crore|lakh|percent|ratio)\b[^)]*\)",
    re.IGNORECASE,
)

_COL_NUMBER_RE = re.compile(r"^\s*\d+\s*$")
_SECTION_MARKER_RE = re.compile(r"^\s*(?:[A-Z]|[IVXLCDM]+)[\.)]?\s*$", re.IGNORECASE)


class ExcelAnalyzer:
    """Analyze Excel (.xlsx) file structure before pandas loading."""

    def _has_meaningful_value(self, value: object) -> bool:
        if value is None:
            return False
        text = str(value).strip()
        return text != "" and text.lower() != "nan"

    def _meaningful_cells(self, row: List[object]) -> List[Tuple[int, object]]:
        return [(c, v) for c, v in enumerate(row[1:], 1) if self._has_meaningful_value(v)]

    def _is_sparse_section_row(self, non_null: List[Tuple[int, object]]) -> bool:
        """
        Detect structural rows like "A | Open ended schemes" that should not
        be classified as real column headers.
        """
        if len(non_null) == 0 or len(non_null) > 3:
            return False

        cols = [c for c, _ in non_null]
        if max(cols) > 3:
            return False

        values = [str(v).strip() for _, v in non_null]
        if not all(values):
            return False

        if len(values) == 1:
            return len(values[0]) < 80 and not _COL_NUMBER_RE.match(values[0])

        first_val = values[0]
        remaining_text = " ".join(values[1:]).strip()
        if _SECTION_MARKER_RE.match(first_val):
            return True

        return len(first_val) <= 6 and not _COL_NUMBER_RE.match(first_val) and len(remaining_text) > 0

    def analyze(self, file_path: str, sheet_name: Optional[str] = None) -> ExcelStructure:
        """
        Analyze the Excel file and return structural metadata.

        The workbook is opened in data_only mode (formulas resolved to values)
        and NOT in read_only mode so we can access merged_cells info.
        """
        logger.info(f"ExcelAnalyzer: analysing {file_path}")

        wb = load_workbook(file_path, read_only=False, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            wb.close()
            raise ValueError(f"No active sheet found in {file_path}")
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        # ── 1. Build a merged-cell lookup ──────────────────────────
        merged_map = self._build_merged_map(ws)

        # ── 2. Read all cell values (with merged-cell fill) ───────
        grid = self._read_grid(ws, max_row, max_col, merged_map)

        # ── 3. Detect title / metadata rows at the top ────────────
        title_hint, unit_hint, first_content_row = self._detect_title_rows(
            grid, max_row, max_col
        )

        # ── 4. Detect header rows ─────────────────────────────────
        header_rows, col_number_row = self._detect_headers(
            grid, first_content_row, max_row, max_col
        )

        # ── 5. Detect data column range ───────────────────────────
        min_data_col, max_data_col = self._detect_col_range(
            grid, header_rows, max_col
        )

        # ── 6. Detect footer rows at the bottom ──────────────────
        data_start_row = (col_number_row or header_rows[-1]) + 1 if header_rows else first_content_row
        footer_start_row, data_end_row = self._detect_footer(
            grid, data_start_row, max_row, max_col
        )

        # ── 7. Resolve merged header values ───────────────────────
        merged_header_values = self._resolve_header_values(
            grid, header_rows, min_data_col, max_data_col
        )

        # Convert header_rows to 0-indexed "skip_rows" count for pandas
        # skip_rows = number of rows before the first header row
        skip_rows = header_rows[0] - 1 if header_rows else 0

        wb.close()

        structure = ExcelStructure(
            skip_rows=skip_rows,
            header_rows=header_rows,
            num_header_rows=len(header_rows),
            col_number_row=col_number_row,
            data_start_row=data_start_row,
            data_end_row=data_end_row,
            footer_start_row=footer_start_row,
            min_data_col=min_data_col,
            max_data_col=max_data_col,
            unit_hint=unit_hint,
            title_hint=title_hint,
            merged_header_values=merged_header_values,
        )

        logger.info(f"ExcelAnalyzer result: {structure.to_dict()}")
        return structure

    # ── Private helpers ───────────────────────────────────────────────

    # ── Regex for "Base Year" / price-type labels in sub-table banners ──
    _BASE_YEAR_RE = re.compile(
        r"base\s*year\s*[:=]?\s*(\d{4}(?:[–\-]\d{2,4})?)", re.IGNORECASE
    )
    _PRICE_LABEL_RE = re.compile(
        r"\b(constant|current)\s+prices?\b", re.IGNORECASE
    )

    # ───────────────────────────────────────────────────────────────────
    #  Whole-workbook analysis (multi-sheet + vertical sub-tables)
    # ───────────────────────────────────────────────────────────────────

    def analyze_workbook(self, file_path: str) -> WorkbookStructure:
        """
        Analyze every sheet in the workbook.  For each sheet, detect
        vertically stacked sub-tables (same columns repeated under
        different banners like "Base Year: 2004-05" or
        "Constant Prices" / "Current Prices").

        Returns a WorkbookStructure with a flat list of SubTable objects
        and a `mergeable` flag indicating whether they can all be
        concatenated into a single DataFrame (same column headers).
        """
        logger.info(f"ExcelAnalyzer.analyze_workbook: {file_path}")
        wb = load_workbook(file_path, read_only=False, data_only=True)
        result = WorkbookStructure(sheet_names=list(wb.sheetnames))

        all_sub_tables: List[SubTable] = []
        workbook_title: Optional[str] = None
        workbook_unit: Optional[str] = None

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            if max_row < 2 or max_col < 2:
                continue

            merged_map = self._build_merged_map(ws)
            grid = self._read_grid(ws, max_row, max_col, merged_map)

            # Detect the title & unit once (from the first sheet that has them)
            title, unit, _ = self._detect_title_rows(grid, max_row, max_col)
            if title and not workbook_title:
                workbook_title = title
            if unit and not workbook_unit:
                workbook_unit = unit

            # Detect sub-tables in this sheet
            subs = self._detect_sub_tables(grid, max_row, max_col, sheet_name)
            # Propagate base_year from first sub-table to siblings in same sheet
            if subs and subs[0].base_year:
                sheet_base_year = subs[0].base_year
                for s in subs[1:]:
                    if not s.base_year:
                        s.base_year = sheet_base_year
            all_sub_tables.extend(subs)

        wb.close()

        result.title_hint = workbook_title
        result.unit_hint = workbook_unit
        result.sub_tables = all_sub_tables

        # Decide mergeability: all sub-tables must have the same header
        # column names (normalised) to be concat-able.
        if len(all_sub_tables) > 1:
            result.mergeable = self._check_mergeable(all_sub_tables)

        logger.info(
            f"ExcelAnalyzer.analyze_workbook: {len(all_sub_tables)} sub-tables "
            f"across {len(wb.sheetnames)} sheets, mergeable={result.mergeable}"
        )
        return result

    # ────────── Sub-table detection within a single sheet ──────────

    def _detect_sub_tables(
        self,
        grid: List[List],
        max_row: int,
        max_col: int,
        sheet_name: str,
    ) -> List[SubTable]:
        """
        Scan a sheet's grid for one or more data tables separated by
        empty-row gaps followed by banner / header rows.

        A "banner" is a short row (1-2 cells) containing text like
        "Base Year : 2004-05" or "Constant Prices" or "(Rupees Crores)".
        """
        sub_tables: List[SubTable] = []
        cursor = 1  # current row being scanned (1-indexed)

        while cursor <= max_row:
            # Skip leading empty rows
            while cursor <= max_row and self._is_empty_row(grid, cursor):
                cursor += 1
            if cursor > max_row:
                break

            # Collect banner metadata (base_year, label, unit) until we
            # hit the first header row (many non-null string cells).
            label: Optional[str] = None
            base_year: Optional[str] = None
            unit: Optional[str] = None
            banner_start = cursor

            while cursor <= max_row:
                row = grid[cursor]
                non_null = self._meaningful_cells(row)
                if len(non_null) == 0:
                    cursor += 1
                    continue

                # Single-cell or double-cell row → banner / metadata
                if len(non_null) <= 2:
                    text = " ".join(str(v) for _, v in non_null).strip()
                    by_m = self._BASE_YEAR_RE.search(text)
                    if by_m:
                        base_year = by_m.group(1)
                    pl_m = self._PRICE_LABEL_RE.search(text)
                    if pl_m:
                        label = pl_m.group(1).title() + " Prices"
                    if _UNIT_PATTERNS.search(text):
                        unit = text.strip("() ")
                    cursor += 1
                    continue

                # Merged cell spanning many columns → all values identical → banner
                unique_vals = set(str(v) for _, v in non_null)
                if len(unique_vals) == 1 and len(non_null) >= 3:
                    text = str(non_null[0][1]).strip()
                    by_m = self._BASE_YEAR_RE.search(text)
                    if by_m:
                        base_year = by_m.group(1)
                    elif not base_year:
                        # Fallback: year in parentheses like "(1999-00)"
                        by_fb = re.search(r"\((\d{4}[–\-]\d{2,4})\)", text)
                        if by_fb:
                            base_year = by_fb.group(1)
                    # Do NOT extract price label from wide merged rows (they are titles)
                    if _UNIT_PATTERNS.search(text):
                        unit = text.strip("() ")
                    cursor += 1
                    continue

                # Header-like row (>=3 non-null, predominantly strings)
                str_count = sum(
                    1 for _, v in non_null
                    if isinstance(v, str) and not _COL_NUMBER_RE.match(str(v).strip())
                )
                if str_count >= 3 or self._is_column_number_row(non_null):
                    break
                # Also break on numeric-dense rows (data rows)
                num_count = sum(
                    1 for _, v in non_null
                    if isinstance(v, (int, float)) and not isinstance(v, bool)
                )
                if num_count >= 3:
                    break
                cursor += 1

            if cursor > max_row:
                break

            # Now detect the header rows + column-number row + data range
            header_rows, col_number_row = self._detect_headers(
                grid, cursor, max_row, max_col
            )
            if not header_rows:
                cursor += 1
                continue

            min_dc, max_dc = self._detect_col_range(grid, header_rows, max_col)
            data_start = (col_number_row or header_rows[-1]) + 1
            # Find data end: scan until an empty-row gap, a new banner, or footer
            data_end = data_start
            for r in range(data_start, max_row + 1):
                row = grid[r]
                non_null = self._meaningful_cells(row)
                if len(non_null) == 0:
                    # Empty row → end of this sub-table's data
                    data_end = r - 1
                    break
                # Check if this row looks like a new banner (1-2 cells, all strings)
                if len(non_null) <= 2:
                    text = " ".join(str(v) for _, v in non_null).strip()
                    if (self._BASE_YEAR_RE.search(text)
                            or self._PRICE_LABEL_RE.search(text)
                            or _UNIT_PATTERNS.search(text)
                            or _FOOTER_PATTERNS.match(text)):
                        data_end = r - 1
                        break
                data_end = r

            # Trim trailing empty rows within data_end
            while data_end >= data_start and self._is_empty_row(grid, data_end):
                data_end -= 1

            if data_end < data_start:
                cursor = data_end + 1
                continue

            merged_hdr = self._resolve_header_values(grid, header_rows, min_dc, max_dc)

            sub = SubTable(
                sheet_name=sheet_name,
                header_rows=header_rows,
                col_number_row=col_number_row,
                data_start_row=data_start,
                data_end_row=data_end,
                min_data_col=min_dc,
                max_data_col=max_dc,
                num_header_rows=len(header_rows),
                merged_header_values=merged_hdr,
                label=label,
                base_year=base_year,
                unit_hint=unit,
            )
            # Check for horizontally-tiled tables (side-by-side with empty separator columns)
            split_subs = self._split_horizontal_sub_tables(grid, sub)
            sub_tables.extend(split_subs)
            for s in split_subs:
                logger.info(
                    f"ExcelAnalyzer: sub-table in [{sheet_name}] "
                    f"rows {s.data_start_row}-{s.data_end_row}, cols {s.min_data_col}-{s.max_data_col}, "
                    f"base_year={s.base_year}, label={s.label}, hsplit={s.horizontal_split}"
                )
            cursor = data_end + 1

        return sub_tables

    def _is_empty_row(self, grid: List[List], r: int) -> bool:
        if r < 1 or r >= len(grid):
            return True
        return all(not self._has_meaningful_value(v) for v in grid[r][1:])

    # ────────── Horizontal sub-table splitting ──────────

    def _split_horizontal_sub_tables(
        self, grid: List[List], sub: SubTable,
    ) -> List[SubTable]:
        """
        Check if a sub-table contains multiple tables placed side-by-side,
        separated by consistently empty columns.  If so, split into one
        SubTable per horizontal block.  Otherwise return [sub] unchanged.
        """
        # 1. Identify columns that are empty in BOTH headers AND data rows
        sample_data_rows = list(range(
            sub.data_start_row,
            min(sub.data_end_row + 1, sub.data_start_row + 10),
        ))
        check_rows = list(sub.header_rows) + sample_data_rows

        empty_cols: set = set()
        for c in range(sub.min_data_col, sub.max_data_col + 1):
            is_empty = True
            for r in check_rows:
                if r < len(grid) and c < len(grid[r]):
                    if self._has_meaningful_value(grid[r][c]):
                        is_empty = False
                        break
            if is_empty:
                empty_cols.add(c)

        if not empty_cols:
            return [sub]

        # 2. Find contiguous non-empty column blocks
        blocks: List[Tuple[int, int]] = []
        block_start: Optional[int] = None
        for c in range(sub.min_data_col, sub.max_data_col + 1):
            if c not in empty_cols:
                if block_start is None:
                    block_start = c
            else:
                if block_start is not None:
                    blocks.append((block_start, c - 1))
                    block_start = None
        if block_start is not None:
            blocks.append((block_start, sub.max_data_col))

        if len(blocks) <= 1:
            return [sub]

        # 3. Reject if block widths are very uneven (not a tiled layout)
        widths = [end - start + 1 for start, end in blocks]
        if max(widths) > 2 * min(widths):
            return [sub]

        # 4. Build a SubTable for each horizontal block
        result: List[SubTable] = []
        for block_start, block_end in blocks:
            # Extract the group label from the first header row
            # (e.g. "BSE", "NSE") — skip generic words like "Period"
            label: Optional[str] = None
            if sub.header_rows:
                first_hdr = sub.header_rows[0]
                for c in range(block_start, block_end + 1):
                    if c < len(grid[first_hdr]):
                        v = grid[first_hdr][c]
                        if (self._has_meaningful_value(v)
                                and isinstance(v, str)
                                and str(v).strip().lower() != "period"):
                            label = re.sub(r'\s+', ' ', str(v).strip())
                            break

            # Build single-level header: prefer bottom header row,
            # fall back to upper rows (but skip the group label text)
            hdr: List[Optional[str]] = []
            for c in range(block_start, block_end + 1):
                val: Optional[str] = None
                for hr in reversed(sub.header_rows):
                    if hr < len(grid) and c < len(grid[hr]):
                        v = grid[hr][c]
                        if self._has_meaningful_value(v):
                            text = re.sub(r'\s+', ' ', str(v).strip())
                            # Skip the label itself (goes to the label column)
                            if label and text.lower() == label.lower():
                                continue
                            val = text
                            break
                hdr.append(val)

            new_sub = SubTable(
                sheet_name=sub.sheet_name,
                header_rows=sub.header_rows,
                col_number_row=sub.col_number_row,
                data_start_row=sub.data_start_row,
                data_end_row=sub.data_end_row,
                min_data_col=block_start,
                max_data_col=block_end,
                num_header_rows=len(sub.header_rows),
                merged_header_values=[hdr],
                label=label,
                base_year=sub.base_year,
                unit_hint=sub.unit_hint,
                horizontal_split=True,
                label_column="source",
            )
            result.append(new_sub)

        # Use the first block's header names as canonical for all blocks
        # to ensure pd.concat alignment (avoids minor formatting diffs)
        if len(result) > 1:
            canonical_hdr = result[0].merged_header_values
            for s in result[1:]:
                s.merged_header_values = canonical_hdr

        logger.info(
            f"ExcelAnalyzer: horizontal split [{sub.sheet_name}] → "
            f"{len(result)} blocks: "
            f"{[(s.min_data_col, s.max_data_col, s.label) for s in result]}"
        )
        return result

    def _normalise_header(self, merged_vals: Optional[List[List[Optional[str]]]]) -> Tuple[str, ...]:
        """Collapse multi-level headers into a normalised tuple for comparison."""
        if not merged_vals:
            return ()
        names = []
        for col_idx in range(len(merged_vals[0])):
            parts = []
            for level in merged_vals:
                if col_idx < len(level) and level[col_idx]:
                    part = re.sub(r"\s+", " ", level[col_idx].strip().lower())
                    if part not in parts:
                        parts.append(part)
            names.append("_".join(parts))
        return tuple(names)

    def _check_mergeable(self, subs: List[SubTable]) -> bool:
        """
        Return True if all sub-tables share the same first column header
        and have ≥50% overlap in column names (allowing for extra columns
        in some sub-tables, e.g. newer states added in later base years).
        """
        if len(subs) <= 1:
            return False
        headers = [self._normalise_header(s.merged_header_values) for s in subs]
        headers = [h for h in headers if h]
        if not headers:
            return False
        # First column must match across all sub-tables
        first_cols = [h[0] for h in headers]
        if len(set(first_cols)) != 1:
            return False
        # Each sub-table must share ≥50% of its columns with at least
        # one other sub-table (handles varying column counts)
        col_sets = [set(h) for h in headers]
        for i, s_i in enumerate(col_sets):
            best_overlap = max(
                len(s_i & s_j) / max(len(s_i), 1)
                for j, s_j in enumerate(col_sets) if j != i
            )
            if best_overlap < 0.5:
                return False
        return True

    def _build_merged_map(self, ws) -> Dict[Tuple[int, int], Tuple[int, int]]:
        """
        Build a dict mapping every cell inside a merged range to the
        top-left anchor cell of that range.
        Key: (row, col) 1-indexed  →  Value: (anchor_row, anchor_col) 1-indexed
        """
        merged_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for merge_range in ws.merged_cells.ranges:
            min_r, min_c = merge_range.min_row, merge_range.min_col
            for r in range(merge_range.min_row, merge_range.max_row + 1):
                for c in range(merge_range.min_col, merge_range.max_col + 1):
                    if (r, c) != (min_r, min_c):
                        merged_map[(r, c)] = (min_r, min_c)
        return merged_map

    def _read_grid(
        self, ws, max_row: int, max_col: int,
        merged_map: Dict[Tuple[int, int], Tuple[int, int]],
    ) -> List[List]:
        """
        Read the worksheet into a 2D list (1-indexed via grid[row][col]).
        Merged cells are filled with the anchor cell's value.
        grid[0] is unused (placeholder) so that grid[1][1] = cell A1.
        """
        # Pre-read all cell values
        raw: Dict[Tuple[int, int], object] = {}
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    raw[(cell.row, cell.column)] = cell.value

        # Build grid with merged-cell propagation
        grid: List[List] = [[None] * (max_col + 1)]  # row 0 placeholder
        for r in range(1, max_row + 1):
            row_vals: List = [None]  # col 0 placeholder
            for c in range(1, max_col + 1):
                if (r, c) in merged_map:
                    anchor = merged_map[(r, c)]
                    row_vals.append(raw.get(anchor))
                else:
                    row_vals.append(raw.get((r, c)))
            grid.append(row_vals)
        return grid

    def _detect_title_rows(
        self, grid: List[List], max_row: int, max_col: int,
    ) -> Tuple[Optional[str], Optional[str], int]:
        """
        Detect title and unit rows at the top of the sheet.
        Returns (title_text, unit_text, first_content_row).

        Title rows: rows with text in only 1-2 columns, or a single merged text
        spanning many columns.  They appear before the real headers.
        """
        title_hint = None
        unit_hint = None
        first_content_row = 1

        for r in range(1, min(max_row + 1, 12)):  # scan first 11 rows
            row = grid[r]
            non_null = self._meaningful_cells(row)

            if len(non_null) == 0:
                # empty row — could be separator between title and headers
                continue

            # Check if this is a unit hint row like "(₹ Crores)"
            if len(non_null) == 1:
                val_str = str(non_null[0][1]).strip()
                if _UNIT_PATTERNS.search(val_str):
                    unit_hint = val_str.strip("() ")
                    continue

            # Check if all non-null values are the same (merged title)
            unique_vals = set(str(v) for _, v in non_null)
            if len(unique_vals) == 1 and len(non_null) >= 2:
                # Merged title row
                if title_hint is None:
                    title_hint = str(non_null[0][1]).strip()
                continue

            # Single value in one cell — likely a title or unit
            if len(non_null) == 1:
                val_str = str(non_null[0][1]).strip()
                # Short single-cell text before headers = likely title/unit
                if len(val_str) < 200 and not _COL_NUMBER_RE.match(val_str):
                    if title_hint is None and not _UNIT_PATTERNS.search(val_str):
                        title_hint = val_str
                    continue

            if self._is_sparse_section_row(non_null):
                if title_hint is None:
                    title_hint = " ".join(str(v).strip() for _, v in non_null)
                continue

            # Multiple non-null values in different columns → likely a header or data row
            # Check if at least 3 columns have values (characteristic of headers)
            if len(non_null) >= 3:
                first_content_row = r
                break

        logger.info(
            f"ExcelAnalyzer: title='{title_hint}', unit='{unit_hint}', "
            f"first_content_row={first_content_row}"
        )
        return title_hint, unit_hint, first_content_row

    def _detect_headers(
        self, grid: List[List], first_content_row: int, max_row: int, max_col: int,
    ) -> Tuple[List[int], Optional[int]]:
        """
        Starting from first_content_row, detect header and column-numbering rows.
        Returns (header_row_numbers, col_number_row_or_None).

        Header rows have many string values across columns.
        Column-numbering rows have sequential integers (1, 2, 3, ...).
        """
        header_rows: List[int] = []
        col_number_row: Optional[int] = None
        consecutive_blank_rows = 0

        for r in range(first_content_row, min(max_row + 1, first_content_row + 10)):
            row = grid[r]
            non_null = self._meaningful_cells(row)
            if len(non_null) == 0:
                consecutive_blank_rows += 1
                if header_rows and consecutive_blank_rows >= 1:
                    break
                continue
            consecutive_blank_rows = 0

            # Check if this is a column-numbering row (1, 2, 3, ...)
            if self._is_column_number_row(non_null):
                col_number_row = r
                continue

            if self._is_sparse_section_row(non_null):
                if header_rows:
                    break
                continue

            # Check if this looks like a header row
            str_count = sum(1 for _, v in non_null if isinstance(v, str) and not _COL_NUMBER_RE.match(str(v).strip()))
            num_count = sum(1 for _, v in non_null if isinstance(v, (int, float)) and not isinstance(v, bool))
            filled_cols = len(non_null)

            # A header row has predominantly string values
            if filled_cols >= 3 and str_count >= max(2, min(3, filled_cols)) and str_count >= num_count:
                header_rows.append(r)
                if len(header_rows) >= 3:
                    break
            elif num_count > str_count and len(header_rows) > 0:
                # We've hit data rows; stop
                break
            elif header_rows:
                break

        if not header_rows:
            header_rows = [first_content_row]

        logger.info(
            f"ExcelAnalyzer: header_rows={header_rows}, "
            f"col_number_row={col_number_row}"
        )
        return header_rows, col_number_row

    def _is_column_number_row(self, non_null: List[Tuple[int, object]]) -> bool:
        """Check if a row contains sequential column numbers like 1, 2, 3, ..."""
        if len(non_null) < 3:
            return False

        numbers = []
        for _, v in non_null:
            try:
                n = int(float(str(v).strip()))
                numbers.append(n)
            except (ValueError, TypeError):
                return False

        if not numbers:
            return False

        if len(set(numbers)) != len(numbers):
            return False
        if numbers[0] != 1:
            return False

        deltas = [curr - prev for prev, curr in zip(numbers, numbers[1:])]
        if any(delta <= 0 for delta in deltas):
            return False

        near_sequential = sum(1 for delta in deltas if delta in (1, 2))
        return near_sequential >= max(len(deltas) - 1, 1)

    def _detect_col_range(
        self, grid: List[List], header_rows: List[int], max_col: int,
    ) -> Tuple[int, int]:
        """Detect the range of columns that contain actual data."""
        min_col = max_col
        max_col_found = 1

        for r in header_rows:
            row = grid[r]
            for c in range(1, len(row)):
                if row[c] is not None:
                    min_col = min(min_col, c)
                    max_col_found = max(max_col_found, c)

        return min_col, max_col_found

    def _detect_footer(
        self, grid: List[List], data_start: int, max_row: int, max_col: int,
    ) -> Tuple[Optional[int], int]:
        """
        Scan from the bottom up to find footer/notes rows.
        Returns (footer_start_row_or_None, last_data_row).
        """
        footer_start = None
        last_data_row = max_row

        for r in range(max_row, max(data_start - 1, 0), -1):
            row = grid[r]
            non_null = self._meaningful_cells(row)

            if len(non_null) == 0:
                # Empty row — could be separator between data and footer
                last_data_row = r - 1
                continue

            # Check if this row matches footer patterns
            first_val = str(non_null[0][1]).strip()
            if _FOOTER_PATTERNS.match(first_val):
                footer_start = r
                last_data_row = r - 1
                continue

            # Check if only 1-2 columns have values and it's text-heavy
            # (characteristic of footnotes in last rows)
            if len(non_null) <= 2 and all(isinstance(v, str) for _, v in non_null):
                text = " ".join(str(v) for _, v in non_null)
                if len(text) > 50:  # Long text in 1-2 columns = likely footnote
                    footer_start = r
                    last_data_row = r - 1
                    continue

            # This row has real data — stop scanning
            break

        # Adjust: if last_data_row still points at an empty row, move up
        while last_data_row > data_start and all(
            not self._has_meaningful_value(v) for v in grid[last_data_row][1:]
        ):
            last_data_row -= 1

        logger.info(
            f"ExcelAnalyzer: footer_start={footer_start}, "
            f"last_data_row={last_data_row}"
        )
        return footer_start, last_data_row

    def _resolve_header_values(
        self, grid: List[List], header_rows: List[int],
        min_col: int, max_col: int,
    ) -> List[List[Optional[str]]]:
        """
        Build a resolved header grid where merged cells are properly filled.
        Returns a list of lists, one per header row, with string values for
        each column in the [min_col, max_col] range.
        """
        if not header_rows:
            return []

        result = []
        previous_row: Optional[List[Optional[str]]] = None
        for r in header_rows:
            row_vals = []
            has_meaningful_header = False
            for idx, c in enumerate(range(min_col, max_col + 1)):
                v = grid[r][c] if c < len(grid[r]) else None
                cell_text = str(v).strip() if self._has_meaningful_value(v) else None
                if cell_text and _COL_NUMBER_RE.match(cell_text):
                    cell_text = None
                if cell_text is None and previous_row and idx < len(previous_row):
                    cell_text = previous_row[idx]
                if cell_text is not None:
                    has_meaningful_header = True
                row_vals.append(cell_text)
            if has_meaningful_header:
                result.append(row_vals)
                previous_row = row_vals
        return result


# Module-level singleton
excel_analyzer = ExcelAnalyzer()
